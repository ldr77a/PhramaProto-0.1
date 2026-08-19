"""후보 조성표를 개별 Excel 통합 문서로 직렬화한다."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from generation.html_formatter import candidate_rows

_STATUS_LABELS = {
    "pass": "배합 가능",
    "warning": "조건부 후보",
    "unresolved": "미해결",
}


def _excel_text(value: str) -> str:
    return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value


def candidate_workbook(candidate) -> bytes:
    """한 후보의 화면 조성표를 실제 .xlsx 바이트로 만든다."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"조성 후보 {candidate.idx}"

    sheet.append([f"조성 후보 {candidate.idx}"])
    sheet.merge_cells("A1:E1")
    sheet["A1"].font = Font(size=16, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.append(["상태", _STATUS_LABELS.get(candidate.status, candidate.status)])
    picks = ", ".join(f"{function}: {name}" for function, name in candidate.pick.items())
    sheet.append(["선정 조합", _excel_text(picks)])
    sheet.append([])
    sheet.append(["성분", "기능", "mg", "%", "근거"])

    for name, function, mg, pct, provenance, _ in candidate_rows(candidate):
        sheet.append(
            [
                _excel_text(name),
                _excel_text(function),
                round(mg, 1) if mg is not None else None,
                round(pct, 2) if pct is not None else None,
                _excel_text(provenance),
            ]
        )

    total_mg = sum(component.mg for component in candidate.components if component.mg)
    total_pct = sum(component.pct for component in candidate.components if component.pct)
    sheet.append(
        [
            "합계",
            None,
            round(total_mg, 1),
            round(total_pct, 2),
            f"총중량 {candidate.total_mg:g}mg",
        ]
    )

    header_fill = PatternFill("solid", fgColor="24313A")
    for cell in sheet[5]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)

    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = f"A5:E{sheet.max_row}"
    for column, width in {"A": 30, "B": 18, "C": 12, "D": 12, "E": 30}.items():
        sheet.column_dimensions[column].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


__all__ = ["candidate_workbook"]
