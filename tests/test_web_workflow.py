from __future__ import annotations

import base64
import json
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from typing import ClassVar

import pytest
from openpyxl import load_workbook

import pharma_proto.app as app_module
from gates.formulation import Component, GateResult
from generation.dose_resolver import DoseResult
from generation.excipient_allocator import Alloc
from generation.generation_loop import Candidate
from pharma_proto.app import create_app, shutdown_app_resources
from pharma_proto.excel_export import candidate_workbook

ROOT = Path(__file__).resolve().parents[1]


class _ElementProbe(HTMLParser):
    _VOID_TAGS: ClassVar[frozenset[str]] = frozenset(
        {
            "area",
            "base",
            "br",
            "col",
            "embed",
            "hr",
            "img",
            "input",
            "link",
            "meta",
            "source",
            "track",
            "wbr",
        }
    )

    def __init__(self) -> None:
        super().__init__()
        self.elements: dict[str, dict[str, object]] = {}
        self._stack: list[str | None] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attributes = dict(attrs)
        element_id = attributes.get("id")
        if element_id is not None:
            self.elements[element_id] = {"attrs": attributes, "text": []}
        if tag not in self._VOID_TAGS:
            self._stack.append(element_id)

    def handle_endtag(self, tag: str) -> None:
        self._stack.pop()

    def handle_data(self, data: str) -> None:
        for element_id in self._stack:
            if element_id is not None:
                text = self.elements[element_id]["text"]
                assert isinstance(text, list)
                text.append(data)

    def attrs(self, element_id: str) -> dict[str, str | None]:
        attrs = self.elements[element_id]["attrs"]
        assert isinstance(attrs, dict)
        return attrs

    def text(self, element_id: str) -> str:
        text = self.elements[element_id]["text"]
        assert isinstance(text, list)
        return " ".join("".join(text).split())


class _FakeLLMService:
    def __init__(self, spec: SimpleNamespace) -> None:
        self._spec = spec

    def parse(self, provider: str, tier: str, api_key: str, question: str):
        return self._spec


@pytest.fixture
def app_factory(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setenv("LOCALAPPDATA", str(tmp_path))
    apps = []

    def build(*, llm_service=None):
        overrides = {
            "SNAPSHOT_PATH": ROOT / "release-data" / "knowledge.sqlite",
            "MANIFEST_PATH": ROOT / "release-data" / "manifest.json",
        }
        if llm_service is not None:
            overrides["LLM_SERVICE"] = llm_service
        app = create_app(overrides)
        app.config["TESTING"] = True
        apps.append(app)
        return app

    yield build

    for app in apps:
        shutdown_app_resources(app)


def _candidate(index: int) -> Candidate:
    api = Component(
        "Acetaminophen",
        role="api",
        mg=500.0,
        pct=83.333,
        function="api",
    )
    diluent = Component(
        "Microcrystalline cellulose",
        role="excipient",
        mg=100.0,
        pct=16.667,
        function="diluent",
    )
    gate = GateResult("게이트1 허용범위", "pass", "허용범위 충족")
    return Candidate(
        idx=index,
        pick={"diluent": "Microcrystalline cellulose"},
        doses=[DoseResult("Acetaminophen", 500.0, "user", "사용자 지정")],
        allocs=[
            Alloc(
                "Microcrystalline cellulose",
                "diluent",
                16.667,
                "filler(q.s.)",
                mg=100.0,
            )
        ],
        components=[api, diluent],
        total_mg=600.0,
        gate_out={"warnings": [], "hard_fails": [], "results": [gate]},
        status="pass",
    )


def test_first_load_only_exposes_api_setup(app_factory) -> None:
    response = app_factory().test_client().get("/")

    assert response.status_code == 200
    probe = _ElementProbe()
    probe.feed(response.get_data(as_text=True))

    assert "hidden" not in probe.attrs("api-setup")
    assert "hidden" in probe.attrs("research-app")
    assert "hidden" in probe.attrs("review-notice")
    assert probe.text("api-cost-notice") == "외부 AI API 호출은 과금 대상입니다."
    assert "연구 검토용 시제품입니다." not in probe.text("api-setup")
    assert "API 설정 변경" in probe.text("research-app")

    catalog = json.loads(probe.text("model-catalog"))
    assert catalog == {
        "openai": {
            "cheap": "gpt-5.6-luna",
            "normal": "gpt-5.6-terra",
            "good": "gpt-5.6-sol",
        },
        "gemini": {
            "cheap": "gemini-3.5-flash-lite",
            "normal": "gemini-3.7-flash",
            "good": "gemini-3.1-pro-preview",
        },
        "claude": {
            "cheap": "claude-haiku-4-5-20251001",
            "normal": "claude-sonnet-5",
            "good": "claude-opus-5",
        },
    }


def test_generate_returns_one_real_xlsx_download_per_candidate(
    app_factory,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    spec = SimpleNamespace(
        apis=[SimpleNamespace(name="Acetaminophen")],
        dosage_form="tablet",
        process="direct compression",
        profile_id="immediate_release_tablet",
    )
    monkeypatch.setattr(
        app_module,
        "run_generation",
        lambda spec, *, repository, offline: [_candidate(i) for i in range(1, 4)],
    )
    client = app_factory(llm_service=_FakeLLMService(spec)).test_client()
    assert client.post(
        "/api/key",
        json={"provider": "openai", "api_key": "test-key"},
    ).status_code == 200

    response = client.post(
        "/api/generate",
        json={
            "provider": "openai",
            "tier": "normal",
            "question": "아세트아미노펜 500 mg 정제",
        },
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["html"].count('class="download-xlsx ') == 3
    assert [item["candidate_idx"] for item in payload["downloads"]] == [1, 2, 3]
    assert [item["filename"] for item in payload["downloads"]] == [
        "조성_후보_1.xlsx",
        "조성_후보_2.xlsx",
        "조성_후보_3.xlsx",
    ]

    workbook_bytes = base64.b64decode(payload["downloads"][0]["content_base64"])
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
    sheet = workbook.active
    assert sheet.title == "조성 후보 1"
    assert sheet["A1"].value == "조성 후보 1"
    assert [sheet.cell(5, column).value for column in range(1, 6)] == [
        "성분",
        "기능",
        "mg",
        "%",
        "근거",
    ]
    assert [sheet.cell(6, column).value for column in range(1, 6)] == [
        "Acetaminophen",
        "API",
        500.0,
        83.33,
        "사용자 지정",
    ]
    assert [sheet.cell(8, column).value for column in range(1, 6)] == [
        "합계",
        None,
        600.0,
        100.0,
        "총중량 600mg",
    ]


@pytest.mark.parametrize(
    "ingredient_name",
    [
        '=HYPERLINK("https://invalid.example")',
        "+1+1",
        "-1+1",
        "@SUM(1,1)",
    ],
)
def test_xlsx_treats_formula_like_ingredient_names_as_text(
    ingredient_name: str,
) -> None:
    candidate = _candidate(1)
    candidate.components[0].name = ingredient_name

    workbook = load_workbook(
        BytesIO(candidate_workbook(candidate)),
        data_only=False,
    )
    ingredient_cell = workbook.active["A6"]

    assert ingredient_cell.data_type == "s"
    assert ingredient_cell.value == f"'{ingredient_name}"
